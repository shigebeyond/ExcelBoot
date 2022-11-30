# -*- coding: utf-8 -*-

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection, colors

wb = Workbook() # 默认生成一个名为Sheet的sheet

# 创建sheet
for name in ['a','b']:
    ws = wb.create_sheet(name)

for sheet in wb:
    sheet.row_dimensions[1].width = 20

for sheet in wb:
    sheet.column_dimensions['c'].width = 40
    sheet.column_dimensions['c'].style = "Hyperlink"
    sheet.column_dimensions['c'].font = Font(name='宋体',color=colors.BLUE,italic=True,size=14)
    sheet.column_dimensions['c'].fill = PatternFill(fill_type='solid', start_color=colors.BLUE)

# 填充数据
ls = [1, 2, 3, 4, 5, 6]
for sheet in wb:
    for i in range(5):
        sheet.append(ls)

for sheet in wb:
    res_row = sheet.row_dimensions.items()
    for i,obj in res_row:
        print(i,obj)
        print('-----------')

res_col = sheet.column_dimensions.items()

for i,obj in res_col:
    print(i,obj)
    print('==========')

wb.save('data/test.xlsx')