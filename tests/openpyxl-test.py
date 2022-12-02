# -*- coding: utf-8 -*-

'''
https://blog.csdn.net/jomes_wang/article/details/111628199
Worksheet 的其他属性:
    title：表格的标题
    max_row：表格的最大行
    min_row：表格的最小行
    max_column：表格的最大列
    min_column：表格的最小列
    rows：按行获取单元格(Cell对象) - 生成器
    columns：按列获取单元格(Cell对象) - 生成器
    values：按行获取表格的内容(数据) - 生成器
'''

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection, colors
from openpyxl.utils.cell import column_index_from_string, get_column_letter

# 列号和字母的互转
print('列号数->列名: ' + get_column_letter(11))
print('列名->列号数: ' + str(column_index_from_string('AD')))

wb = Workbook() # 默认生成一个名为Sheet的sheet

# 创建sheet
for name in ['a','b']:
    ws = wb.create_sheet(name)

# 修改行
for sheet in wb:
    # sheet.row_dimensions[1].width = 140 # 无用
    sheet.row_dimensions[2].height = 40

# 修改列
for sheet in wb:
    sheet.column_dimensions['c'].width = 40
    # sheet.column_dimensions['c'].height = 40 # 无用
    sheet.column_dimensions['c'].font = Font(name='宋体',color=colors.BLUE,italic=True,size=14)
    sheet.column_dimensions['c'].fill = PatternFill(fill_type='solid', start_color=colors.BLUE)

# 填充数据
ls = [1, 2, 3, 4, 5, 6]
for sheet in wb:
    for i in range(5):
        row = [sheet.title + str(it*i) for it in ls]
        sheet.append(row)

print("输出行dim")
for sheet in wb:
    res_row = sheet.row_dimensions.items()
    for i,obj in res_row:
        print(i,obj)

print("输出列dim")
res_col = sheet.column_dimensions.items()
for i,obj in res_col:
    print(i,obj)

sheet = wb['a']
print("输出行全部值")
print(list(sheet.values))
i = 0
for items in sheet["A1:C4"]: # 输出顺序：先逐行，后逐列
    i += 1
    print(f"A1:C3-第{i}行")
    print([cell.value for cell in items])

# 修改列
col = sheet['A']

# 修改单元格
cell = sheet['A1']
cell.style = "Hyperlink"
cell.value = 'test'
# cell.width = 40 # AttributeError: 'Cell' object has no attribute 'width'

# openpyxl 样式
font = Font(name="微软雅黑", size=10, bold=True, italic=False, color="FF0000")
alignment = Alignment(horizontal="center", vertical="center")
for i in range(1, 10):
    cell = ws.cell(row=1, column=i)
    print(cell.value)
    cell.font = font
    cell.alignment = alignment

# 保存
wb.save('data/test.xlsx')