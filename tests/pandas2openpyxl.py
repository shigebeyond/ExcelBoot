# Pandas与openpyxl库 结合
# https://blog.csdn.net/weixin_41261833/article/details/120169556

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook

# df 转 openpyxl
def df2open():
    print('df 转 openpyxl')
    data = {
        "姓名": ["张三", "李四"],
        "性别": ["男", "女"],
        "年龄": [15, 25],
    }
    df = pd.DataFrame(data)
    print(df)
    wb = Workbook()
    ws = wb.active
    # df 转 openpyxl
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    wb.save("../data/pandas.xlsx")

# openpyxl 转 df
def open2df():
    print('openpyxl 转 df')
    workbook = load_workbook(filename="../data/pandas.xlsx")
    sheet = workbook.active
    values = sheet.values
    df = pd.DataFrame(values)
    print(df)

df2open()
open2df()