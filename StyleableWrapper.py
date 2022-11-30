#!/usr/bin/python3
# -*- coding: utf-8 -*-

from openpyxl.styles.styleable import StyleableObject
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection
import color

'''
# StyleableObject包装器, 包装Cell/ColumnDimension/RowDimension等对象, 主要是给这些对象设置样式
https://blog.csdn.net/qq_44614026/article/details/109707265
'''
class StyleableWrapper(object):

    def __init__(self, obj: StyleableObject):
        self.obj = obj # StyleableObject 对象, 如 Cell/ColumnDimension/RowDimension

    # 设置宽度
    def width(self, w):
        self.obj.width = w

    # 设置高度
    def height(self, h):
        self.obj.height = h

    # 设置样式
    def style(self, s):
        self.obj.style = s

    def font(self, config):
        '''
        设置字体
        :param config name=字体类型, size=字体大小, bold=是否加粗, italic=是否斜体, color=字体颜色
        :return:
        '''
        # self.obj.font = Font(name='宋体', color='FFFF00', bold=True, italic=True, size=14)
        if 'color' in config:
            config['color'] = color.get_rgb(config['color']) # 修正颜色
        self.obj.font = Font(**config)

    def alignment(self, config):
        '''
        设置对齐样式
        :param config horizontal=水平对齐模式, vertical=垂直对齐模式, text_rotation=旋转角度, wrap_text=是否自动换行
        :return:
        '''
        # self.obj.alignment = Alignment(horizontal="center", vertical="center", text_rotation=45, wrap_text=True)
        self.obj.alignment = Alignment(**config)

    def border(self, config):
        '''
        设置边框样式
        :param config color=边线颜色, style=边线样式(double/mediumDashDotDot/slantDashDot/dashDotDot/dotted/hair/mediumDashed/dashed/dashDot/thin/mediumDashDot/medium/thick)
        :return:
        '''
        # side = Side(style="thick", color="FFFF0000")
        if 'color' in config:
            config['color'] = color.get_rgb(config['color']) # 修正颜色
        side = Side(**config)
        self.obj.border = Border(left=side, right=side, top=side, bottom=side)


    # 填充颜色
    def fill(self, color):
        self.obj.fill = PatternFill(fill_type='solid', start_color=color)
