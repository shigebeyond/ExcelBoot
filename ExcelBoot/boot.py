#!/usr/bin/python3
# -*- coding: utf-8 -*-

import fnmatch
import os
from pyutilb.util import *
from pyutilb import log, YamlBoot, BreakException
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils.cell import column_index_from_string, get_column_letter, range_boundaries
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.drawing.image import Image
from ExcelBoot.db import Db
from ExcelBoot.styleable_wrapper import StyleableWrapper
from ExcelBoot.df_col_mapper import DfColMapper
import ExcelBoot.plt_ext as plt
import platform
is_win = platform.system().lower() == 'windows'
if is_win:
    import pythoncom
    pythoncom.CoInitialize()
    import win32com.client

# excel操作的基于yaml的启动器
class Boot(YamlBoot):

    def __init__(self):
        super().__init__()
        # 动作映射函数
        actions = {
            'start_edit': self.start_edit,
            'end_edit': self.end_edit,
            'switch_sheet': self.switch_sheet,
            'connect_db': self.connect_db,
            'query_db': self.query_db,
            'export_df': self.export_df,
            'export_db': self.export_db,
            'map_df_cols': self.map_df_cols,
            'map_cols': self.map_cols,
            'set_cell_value': self.set_cell_value,
            'cells': self.cells,
            'cols': self.cols,
            'rows': self.rows,
            'insert_rows': self.insert_rows,
            'delete_cols': self.delete_cols,
            'delete_rows': self.delete_rows,
            'merge_cells': self.merge_cells,
            'unmerge_cells': self.unmerge_cells,
            'insert_image': self.insert_image,
            'insert_file': self.insert_file,
            'insert_plot': self.insert_plot,
        }
        self.add_actions(actions)

        # excel对象
        self.wb: Workbook = None # book
        self.ws: Worksheet = None # sheet
        self.sheet = None # sheet名

    # --------- 动作处理的函数 --------
    # 开始编辑excel
    def start_edit(self, file):
        if self.wb != None:
            raise Exception(f"尚有excel文件[{self.file}]在编辑中, 没有结束编辑")

        self.file = replace_var(file)
        self.reload_wb()

    # 结束编辑excel -- 保存
    def end_edit(self, _):
        if self.wb == None:
            return
        self.wb.save(self.file)
        self.wb.close()
        self.wb = None
        self.ws = None

    # 切换sheet
    # https://blog.csdn.net/JunChen681/article/details/126053045
    def switch_sheet(self, sheet):
        self.sheet = replace_var(sheet)
        self.reload_ws()

    # 重载Workbook
    def reload_wb(self):
        '''
        if os.path.isfile(self.file):
            self.wb = load_workbook(self.file)
        else:
            self.wb = Workbook()
        '''
        # 创建Workbook
        self.wb = Workbook()
        # 删除默认创建的sheet
        sheets = self.wb.sheetnames
        if len(sheets) > 0:
            del self.wb[sheets[0]]

        self.sheet = None
        self.ws = None

    # 重载Worksheet
    def reload_ws(self):
        if self.sheet in self.wb.sheetnames: # 有则返回
            self.ws = self.wb[self.sheet]
        else: # 无则创建
            self.ws = self.wb.create_sheet(self.sheet)

    # 连接db
    def connect_db(self, config):
        self.db = Db(config['ip'], config['port'], config['dbname'], config['user'], config['password'], config['echo_sql'])

    # 查询db
    def query_db(self, config):
        for var, sql in config.items():
            sql = replace_var(sql)
            # 查询
            df = self.db.query_dataFrame(sql)
            set_var(var, df)

    # 导出 DataFrame 数据
    def export_df(self, config):
        if isinstance(config, str):
            var_df = config
            config = {}
        else:
            var_df = config["df"]
        # 获得导出的变量
        df = self.get_var_DataFrame(var_df)
        # 导出
        self.do_export(df, var_df)

    # 导出 sql 数据
    def export_db(self, sql):
        sql = replace_var(sql)
        # 查询
        df = self.db.query_dataFrame(sql)
        # 导出 df
        self.do_export(df, sql)

    # 真正的导出 df
    def do_export(self, df, var):
        if 'select' in var or 'SELECT' in var:
            type = "select sql"
        else:
            type = "列表变量"
        if len(df) == 0:
            log.debug(f"{type}[{var}]为空, 不用导出excel")
            return

        # df转sheet
        self.df2sheet(df)

    # 获得指定变量, 并保证是DataFrame类型
    def get_var_DataFrame(self, var):
        val = get_var(var)
        # 检查类型
        if not isinstance(val, pd.DataFrame):
            if not isinstance(val, (list, tuple, set)):
                raise Exception(f"变量[{var}]值不是DataFrame或list或tuple或set: {val}")
            # list转DataFrame
            if len(val) == 0: # 空
                val = pd.DataFrame()
            else:
                fields = val[0].keys()
                val = pd.DataFrame(val, columns=fields)
            # 回写变量
            set_var(var, val)
        return val

    # df转sheet
    def df2sheet(self, df):
        if df.empty:
            return

        # 1 df转list: dataframe_to_rows()
        rows = dataframe_to_rows(df, index=False, header=True)

        # 2 写sheet
        # 2.1 sheet为空：直接插
        if self.ws._current_row == 0:
            for row in rows:
                self.ws.append(row)
            return

        # 2.2 sheet不空：改数据
        rows = list(rows) # generater 转 list
        r = len(rows)
        c = len(rows[0])
        bound = self.check_bound(f'1,1:{r},{c}')
        self.do_set_cell_values(bound, rows)

    # sheet转df
    def sheet2df(self, has_header):
        # 获得sheet数据
        values = self.ws.values  # generator
        values = list(values)  # 转list
        # 空
        if len(values) == 0:
            return pd.DataFrame()

        # 列名
        if has_header:
            # 第一行作为列名
            # columns = self.iterate_cell_values('1')
            columns = values[0]
            # 删除第一行
            del values[0]
        else:  # ABC作为列名
            columns = [get_column_letter(i) for i in range(1, len(values[0]) + 1)]

        # 转df
        return pd.DataFrame(values, columns=columns)

    # 列表变量变换
    # :param clos 变量名:变换函数
    def map(self, cols):
        # 构建df
        df = pd.DataFrame()
        # 逐列转换
        for col, expr in cols.items():
            # 列即变量
            df['it'] = get_var(col)
            # 构建df列变换器
            mapper = DfColMapper(df)
            mapper.map(col, expr)
            # 回写变量
            set_var(col, df['it'])

    # df列变换
    def map_df_cols(self, cols, var_df):
        # 获得df
        df = self.get_var_DataFrame(var_df)
        self.do_map_df_cols(cols, df)

    # 真正的df列变换
    def do_map_df_cols(self, cols, df):
        # 构建df列变换器
        mapper = DfColMapper(df)
        # 逐列转换
        for col, expr in cols.items():
            mapper.map(col, expr)

    # sheet中列变换
    # 实现是先将sheet转为df，用df来做列变化
    def map_cols(self, cols):
        # 检查是否配置了 header，标识是否有表头(即列名)
        has_header = False
        if 'header' in cols:
            has_header = cols['header']
            del cols['header']

        # 1 sheet转df
        df = self.sheet2df(has_header)

        # 2 df转换
        self.do_map_df_cols(cols, df)

        # 3 df转sheet
        self.df2sheet(df)

    # 设置单元格的值
    def set_cell_value(self, config):
        # 遍历每个范围来设置值
        for bound, value in config.items():
            # 范围+值都要替换变量
            bound = self.check_bound(bound)
            if isinstance(value, str):
                value = replace_var(value)

            # 1 设置多值
            if isinstance(value, (list, tuple, set, pd.Series)):
                self.do_set_cell_values(bound, value)
                return

            # 2 设置单值
            for cell in self.iterate_cells(bound):
                cell.value = value

    def do_set_cell_values(self, bound, vals):
        '''
        根据范围类型来设置多个单元格的值， 参考 iterate_cells() 的实现
        :param bound: 范围
        :param vals: 多个值，其维度(1维或2维)与范围中的行列对应
        '''
        # 1 范围 ws["A1:C3"], ws["A:C"], ws[1:3]
        if ':' in bound:
            # 输出顺序：先逐行，后逐列
            r = 0
            for row in self.ws[bound]:
                c = 0
                for cell in row:
                    cell.value = vals[r][c]
                    c += 1
                r += 1
            return

        # 2 单个单元格 ws["A1"]
        mat = re.match(r'\w+\d+', bound)  # 匹配: 字母+数字
        if mat != None:
            self.ws[bound].value = vals
            return

        # 3 单列或单行 ws["A"], ws[1]
        i = 0
        for cell in self.ws[bound]:
            i += 1
            cell.value = vals[i]

    # 读取单元格的值
    def get_cell_value(self, config):
        # 遍历每个范围来读取值
        for var, bound in config.items():
            # 范围要替换变量
            bound = self.check_bound(bound)
            # 读取范围单元格的值
            value = self.do_get_cell_value(bound)
            # 设置变量
            set_var(var, value)

    def do_get_cell_value(self, bound):
        '''
        根据范围类型来读取多个单元格的值， 参考 iterate_cells() 的实现
        :param bound: 范围
        :return 多个值，其维度(1维或2维)与范围中的行列对应
        '''

        # 1 范围 ws["A1:C3"], ws["A:C"], ws[1:3]
        if ':' in bound:
            # 输出顺序：先逐行，后逐列
            # 以下2种实现是一样的
            # r = []
            # for row in self.ws[bound]:
            #     vs = [cell.value for cell in row]
            #     r.append(vs)
            # return r
            # 参考 Worksheet.__getitem__() 的实现
            min_col, min_row, max_col, max_row = range_boundaries(bound)
            return tuple(self.ws.iter_rows(min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col, values_only=True))

        # 2 单个单元格 ws["A1"]
        mat = re.match(r'\w+\d+', bound)  # 匹配: 字母+数字
        if mat != None:
            return self.ws[bound].value

        # 3 单列或单行 ws["A"], ws[1]
        return [cell.value for cell in self.ws[bound]]

    # 循环rows, 如 rows(1:3)
    # :param styles 每个迭代中要应用的样式
    # :param bound 范围
    def rows(self, styles, bound):
        self.do_for_styleable(styles, self.iterate_rows(bound), f"rows({bound})")

    # 循环cols, 如 cols(A:B)
    # :param styles 每个迭代中要应用的样式
    # :param bound 范围
    def cols(self, styles, bound):
        self.do_for_styleable(styles, self.iterate_cols(bound), f"cols({bound})")

    # 循环cells, 如 cells(A1:B3)
    # :param styles 每个迭代中要应用的样式
    # :param bound 范围
    def cells(self, styles, bound):
        self.do_for_styleable(styles, self.iterate_cells(bound), f"cells({bound})")

    # 循环对象应用样式
    def do_for_styleable(self, styles, objs, label):
        log.debug(f"-- Loop start: {label} -- ")
        for obj in objs:
            StyleableWrapper(obj).use_styles(styles)  # 应用样式
        log.debug(f"-- Loop finish: {label} -- ")

    # 迭代指定范围内的行
    def iterate_rows(self, bound):
        bound = self.check_bound(bound)
        # return map(lambda row: self.ws.row_dimensions[row], self.build_row_range(bound))
        for row in self.build_row_range(bound):
            yield self.ws.row_dimensions[row]

    # 迭代指定范围内的列
    def iterate_cols(self, bound):
        bound = self.check_bound(bound)
        for col in self.build_col_range(bound):
            yield self.ws.column_dimensions[col]

    # 构建行范围
    # :param bound 如 1:3
    def build_row_range(self, bound):
        mat1 = re.match(r'\d+', bound)
        mat2 = re.match(r'\d+:\d+', bound)
        if mat1 == None and mat2 == None:
            raise Exception("无效行范围: " + range)

        # 1 两值
        if ':' in bound:
            # 分割开始值+结束值
            start, end = bound.split(':')
            return range(int(start), int(end) + 1)

        # 2 单值
        return [int(bound)]

    # 构建列范围
    # :param bound 如 A:B
    def build_col_range(self, bound):
        bound = bound.upper() # 转大写
        mat1 = re.match(r'\w+', bound)
        mat2 = re.match(r'\w+:\w+', bound)
        if mat1 == None and mat2 == None:
            raise Exception("无效列范围: " + range)

        # 1 两值
        if ':' in bound:
            # 分割开始值+结束值
            start, end = bound.split(':')
            # 获得开始索引+结束索引
            start = column_index_from_string(start) # 列名 转 列号数
            end = column_index_from_string(end)
            # 构建yield迭代器
            for i in range(start, end + 1):
                yield get_column_letter(i) # 列号数 转 列名
            return

        # 2 单值
        return [bound]

    # 迭代指定范围内的单元格的值
    def iterate_cell_values(self, bound):
        bound = self.check_bound(bound)
        for cell in self.iterate_cells(bound):
            yield cell.value

    def iterate_cells(self, bound):
        '''
        迭代指定范围内的单元格
        https://blog.csdn.net/weixin_48668114/article/details/126444151

        :param bound: 区域 ws["A1:C3"], ws["A:C"], ws[1:3]
                      单行 ws["1"]
                      单列 ws["A"]
                      单元格 ws["A1"]
        :return:
        '''
        # 1 范围 ws["A1:C3"], ws["A:C"], ws[1:3]
        if ':' in bound:
            # 输出顺序：先逐行，后逐列
            for row in self.ws[bound]:
                for cell in row:
                    yield cell
            return

        # 2 单个单元格 ws["A1"]
        mat = re.match(r'\w+\d+', bound) # 匹配: 字母+数字
        if mat != None:
            yield self.ws[bound]
            return

        # 3 单列或单行 ws["A"], ws[1]
        for cell in self.ws[bound]:
            yield cell

    # 检查范围+替换变量+转换范围格式
    def check_bound(self, bound):
        # 1 简单检查范围格式: 字母大小写都可以
        if re.match(r'[\w+\d:,]+', bound) == None:  # 匹配： 字母数字:
            raise Exception('无效范围: ' + bound)

        # 2 替换变量
        bound = replace_var(bound)

        # 3 转换范围"起始行,起始列:结束行,结束列"为openpyxl格式，就是行在前变为列在前，如 1,2:3,4 转 B1:D3
        if ',' in bound:
            def replace(match) -> str:
                r1 = match.group(1)
                c1 = get_column_letter(int(match.group(2)))
                r2 = match.group(3)
                c2 = get_column_letter(int(match.group(4)))
                # 列放到行前面
                return f"{c1}{r1}:{c2}{r2}"
            return re.sub(rf'(\d+),(\d+):(\d+),(\d)', replace, bound)

        return bound

    # 插入列
    # :param param 起始列,插入列数
    def insert_cols(self, param):
        idx, amount = param.split(',')
        self.ws.insert_cols(idx, amount)

    # 插入行
    # :param param 起始行,插入行数
    def insert_rows(self, param):
        idx, amount = param.split(',')
        self.ws.insert_rows(int(idx), int(amount))

    # 删除列
    # :param param 起始列,删除列数
    def delete_cols(self, param):
        idx, amount = param.split(',')
        self.ws.delete_cols(idx, amount)

    # 删除行
    # :param param 起始行,删除行数
    def delete_rows(self, param):
        idx, amount = param.split(',')
        self.ws.delete_rows(idx, amount)

    # 合并单元格
    # https://blog.csdn.net/ovejur/article/details/123982122
    def merge_cells(self, param):
        # self.wb.merge_cells(start_row=7, start_column=1, end_row=8, end_column=3)
        if re.match(r'\d+,\d+:\d+,\d+', param):
            start_row, start_column, end_row, end_column = param.replace(':', ',').split(',')
            self.wb.merge_cells(start_row=start_row, start_column=start_column, end_row=end_row, end_column=end_column)
            return

        # self.wb.merge_cells("C1:D2")
        return self.wb.merge_cells(param)

    # 取消合并单元格
    def unmerge_cells(self, param):
        # self.wb.unmerge_cells(start_row=7, start_column=1, end_row=8, end_column=3)
        if re.match(r'\d+,\d+:\d+,\d+', param):
            start_row, start_column, end_row, end_column = param.replace(':', ',').split(',')
            self.wb.unmerge_cells(start_row=start_row, start_column=start_column, end_row=end_row, end_column=end_column)
            return

        # self.wb.unmerge_cells("C1:D2")
        return self.wb.unmerge_cells(param)

    # 插入图片
    def insert_image(self, config):
        for bound, opt in config.items():
            if isinstance(opt, str):
                file = opt
                size = None
            else:
                file = opt['image']
                size = opt['size'].split(',')
            # 添加图片
            img = Image(file)
            if size != None:
                img.width = int(size[0])
                img.height = int(size[1])
            self.ws.add_image(img, bound)

    # 插入文件
    def insert_file(self, config):
        if not is_win:
            raise Exception(f'由于 insert_file() 使用的是 pywin32 库, 非 windows 系统不能使用')

        # 由于要使用不同的库，因此先保存openpyxl
        self.end_edit()

        # 使用 pywin32 库，来插入sql附件
        # 报错 https://blog.csdn.net/dantegarden/article/details/77547524
        # 成功 https://blog.csdn.net/weixin_39727402/article/details/110021522
        # app = win32com.client.Dispatch('Excel.Application')
        app = win32com.client.gencache.EnsureDispatch('Excel.Application')
        # app = Dispatch('kwps.application')
        app.Visible = True  # 显式打开excel 调试设置True

        # 打开excel文件
        wb = app.Workbooks.Open(self.file)
        try:
            ws = wb.Sheets(self.sheet)
            for bound, file in config.items():
                # 插入附件 To assign an object for OLEObject(=EMBED("Packager Shell Object","")).
                # obj = ws.Shapes.AddOLEObject(ClassType='Paint.Picture', Filename=file, Link=False)  # 报错: err in AddOLEObject: pywintypes.com_error: (-2147352567, '发生意外。', (0, None, None, None, 0, -2146827284), None)
                obj = ws.OLEObjects().Add(ClassType=None, Filename=file, Link=False, DisplayAsIcon=True, Width=18, Height=50)  # 成功

                # 定位附件到指定单元格
                bound = self.check_bound(bound) # 范围要替换变量
                min_col, min_row, _, _ = range_boundaries(bound)
                cell = ws.Cells(min_col, min_row)
                obj.Left = cell.Left
                obj.Top = cell.Top
        except Exception as e:
            print(f"插入文件[{file}]报错: {e}")
        finally:
            wb.SaveAs(self.file)
            app.Quit()

            # 重新打开openpyxl
            self.reload_wb()
            self.reload_ws()

    # 插入plot绘图
    def insert_plot(self, config):
        for bound, opt in config.items():
            # 获得df的变量
            var_df = opt["df"] # df变量名
            del opt["df"]
            df = self.get_var_DataFrame(var_df) # df
            # 绘图
            file = plt.plot(df, **opt)
            # 添加图片
            img = Image(file)
            self.ws.add_image(img, bound)

# cli入口
def main():
    # 基于yaml的执行器
    boot = Boot()
    # 读元数据：author/version/description
    dir = os.path.dirname(__file__)
    meta = read_init_file_meta(dir + os.sep + '__init__.py')
    # 步骤配置的yaml
    step_files = parse_cmd('ExcelBoot', meta['version'])
    if len(step_files) == 0:
        raise Exception("Miss step config file or directory")
    try:
        # 执行yaml配置的步骤
        boot.run(step_files)
    except Exception as ex:
        log.error(f"Exception occurs: current step file is {boot.step_file}", exc_info = ex)
        raise ex

if __name__ == '__main__':
    main()