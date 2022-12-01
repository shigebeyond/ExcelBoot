#!/usr/bin/python3
# -*- coding: utf-8 -*-

import time
import sys
import os
import fnmatch
from pathlib import Path
from pyutilb.util import *
from pyutilb import log, ocr_youdao
import ast
import pandas as pd
from db import Db
from shutil import copyfile
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection
import platform
from styleable_wrapper import StyleableWrapper
from df_col_mapper import DfColMapper

# 跳出循环的异常
class BreakException(Exception):
    def __init__(self, condition):
        self.condition = condition # 跳转条件

# excel操作的基于yaml的启动器
class Boot(object):

    def __init__(self):
        # 步骤文件所在的目录
        self.step_dir = None
        # 已下载过的url对应的文件，key是url，value是文件
        self.downloaded_files = {}
        # 动作映射函数
        self.actions = {
            'sleep': self.sleep,
            'print': self.print,
            'for': self.do_for,
            'once': self.once,
            'break_if': self.break_if,
            'moveon_if': self.moveon_if,
            'moveon_if_exist_by': self.moveon_if_exist_by,
            'break_if_exist_by': self.break_if_exist_by,
            'break_if_not_exist_by': self.break_if_not_exist_by,
            'include': self.include,
            'set_vars': self.set_vars,
            'print_vars': self.print_vars,
            'start_edit': self.start_edit,
            'end_edit': self.end_edit,
            'switch_sheet': self.switch_sheet,
            'connect_db': self.connect_db,
            'query_db': self.query_db,
            'export_excel': self.export_excel,
            'map_cols': self.map_cols,
            'cells': self.cells,
            'cols': self.cols,
            'rows': self.rows,
            'insert_rows': self.insert_rows,
            'remove_cols': self.remove_cols,
            'remove_rows': self.remove_rows,
            'merge_cells': self.merge_cells,
            'unmerge_cells': self.unmerge_cells,
        }
        set_var('boot', self)
        # 当前文件
        self.step_file = None
        self.wb = None # book
        self.ws = None # sheet
        self.sheet = None # sheet名

    '''
    执行入口
    :param step_files 步骤配置文件或目录的列表
    '''
    def run(self, step_files):
        for path in step_files:
            # 1 模式文件
            if '*' in path:
                dir, pattern = path.rsplit(os.sep, 1)  # 从后面分割，分割为目录+模式
                if not os.path.exists(dir):
                    raise Exception(f'Step config directory not exist: {dir}')
                self.run_1dir(dir, pattern)
                return

            # 2 不存在
            if not os.path.exists(path):
                raise Exception(f'Step config file or directory not exist: {path}')

            # 3 目录: 遍历执行子文件
            if os.path.isdir(path):
                self.run_1dir(path)
                return

            # 4 纯文件
            self.run_1file(path)

    # 执行单个步骤目录: 遍历执行子文件
    # :param path 目录
    # :param pattern 文件名模式
    def run_1dir(self, dir, pattern ='*.yml'):
        # 遍历目录: https://blog.csdn.net/allway2/article/details/124176562
        files = os.listdir(dir)
        files.sort() # 按文件名排序
        for file in files:
            if fnmatch.fnmatch(file, pattern): # 匹配文件名模式
                file = os.path.join(dir, file)
                if os.path.isfile(file):
                    self.run_1file(file)

    # 执行单个步骤文件
    # :param step_file 步骤配置文件路径
    # :param include 是否inlude动作触发
    def run_1file(self, step_file, include = False):
        # 获得步骤文件的绝对路径
        if include: # 补上绝对路径
            if not os.path.isabs(step_file):
                step_file = self.step_dir + os.sep + step_file
        else: # 记录目录
            step_file = os.path.abspath(step_file)
            self.step_dir = os.path.dirname(step_file)

        log.debug(f"Load and run step file: {step_file}")
        # 获得步骤
        steps = read_yaml(step_file)
        self.step_file = step_file
        # 执行多个步骤
        self.run_steps(steps)

    # 执行多个步骤
    def run_steps(self, steps):
        # 逐个步骤调用多个动作
        for step in steps:
            for action, param in step.items():
                self.run_action(action, param)

    '''
    执行单个动作：就是调用动作名对应的函数
    :param action 动作名
    :param param 参数
    '''
    def run_action(self, action, param):
        log.debug(f"handle action: {action}={param}")

        is_for_action = '(' in action # 是否循环动作
        if is_for_action: # 解析循环动作
            action, params = parse_func(action)
            n = params[0]

        if action not in self.actions:
            raise Exception(f'Invalid action: [{action}]')

        # 调用动作对应的函数
        func = self.actions[action]
        if is_for_action: # 循环动作: 多加了个参数-循环变量n
            func(param, n)
        else:
            func(param)

    # --------- 动作处理的函数 --------
    # 当前url
    # @property
    # def current_url(self):
    #     if self.driver == None:
    #         return None
    #     return self.driver.current_url

    # 解析动作名for(n)中的n: 或数字或列表
    def parse_for_n(self, n):
        if n == None or n == '':
            return None

        # 1 数字
        if n.isdigit():
            return int(n)

        # 2 变量表达式, 必须是int/list/df.Series类型
        expr = "${" + n + "}"
        n = replace_var(expr, False)

        # fix bug: pd.Series == None 居然返回pd.Series
        if isinstance(n, pd.Series):
            return n
        if n == None or not (isinstance(n, list) or isinstance(n, int)):
            raise Exception(f'Variable in for({n}) parentheses must be int or list or pd.Series type')
        return n

    # for循环
    # :param steps 每个迭代中要执行的步骤
    # :param n 循环次数/循环列表变量名
    def do_for(self, steps, n = None):
        n = self.parse_for_n(n)
        label = f"for({n})"
        # 循环次数
        # fix bug: pd.Series == None 居然返回pd.Series
        n_null = (not isinstance(n, pd.Series)) and n == None
        if n_null:
            n = sys.maxsize # 最大int，等于无限循环次数
            label = f"for(∞)"
        # 循环的列表值
        items = None
        if isinstance(n, list) or isinstance(n, pd.Series):
            items = n
            n = len(items)
        log.debug(f"-- Loop start: {label} -- ")
        last_i = get_var('for_i', False) # 旧的索引
        last_v = get_var('for_v', False) # 旧的元素
        try:
            for i in range(n):
                # i+1表示迭代次数比较容易理解
                log.debug(f"{i+1}th iteration")
                set_var('for_i', i+1) # 更新索引
                if n_null:
                    v = None
                else:
                    v = items[i]
                set_var('for_v', v) # 更新元素
                self.run_steps(steps)
        except BreakException as e:  # 跳出循环
            log.debug(f"-- Loop break: {label}, break condition: {e.condition} -- ")
        else:
            log.debug(f"-- Loop finish: {label} -- ")
        finally:
            set_var('for_i', last_i) # 恢复索引
            set_var('for_v', last_v) # 恢复元素

    # 执行一次子步骤，相当于 for(1)
    def once(self, steps):
        self.do_for(steps, 1)

    # 检查并继续for循环
    def moveon_if(self, expr):
        # break_if(条件取反)
        self.break_if(f"not ({expr})")

    # 跳出for循环
    def break_if(self, expr):
        val = eval(expr, globals(), bvars)  # 丢失本地与全局变量, 如引用不了json模块
        if bool(val):
            raise BreakException(expr)

    # 检查并继续for循环
    def moveon_if_exist_by(self, config):
        self.break_if_not_exist_by(config)

    # 跳出for循环
    def break_if_not_exist_by(self, config):
        if not self.exist_by_any(config):
            raise BreakException(config)

    # 跳出for循环
    def break_if_exist_by(self, config):
        if self.exist_by_any(config):
            raise BreakException(config)

    # 加载并执行其他步骤文件
    def include(self, step_file):
        self.run_1file(step_file, True)

    # 设置变量
    def set_vars(self, vars):
        for k, v in vars.items():
            v = replace_var(v)  # 替换变量
            set_var(k, v)

    # 打印变量
    def print_vars(self, _):
        log.info(f"Variables: {bvars}")

    # 睡眠
    def sleep(self, seconds):
        seconds = replace_var(seconds)  # 替换变量
        time.sleep(int(seconds))

    # 打印
    def print(self, msg):
        msg = replace_var(msg)  # 替换变量
        log.debug(msg)

    # 开始编辑excel
    def start_edit(self, file):
        self.file = replace_var(file)
        self.reload_wb()

    # 结束编辑excel -- 保存
    def end_edit(self, _):
        self.wb.save(self.file)
        self.wb.close()
        self.wb = None
        self.ws = None

    # 切换sheet
    # https://blog.csdn.net/JunChen681/article/details/126053045
    def switch_sheet(self, sheet):
        self.sheet = replace_var(sheet)
        self.reload_ws()

    # 重载Workbook
    def reload_wb(self):
        self.wb = Workbook()
        '''
        if os.path.isfile(self.file):
            self.wb = load_workbook(self.file)
        else:
            self.wb = Workbook()
        '''
        self.sheet = None
        self.ws = None

    # 重载Worksheet
    def reload_ws(self):
        if self.sheet not in self.wb.sheetnames:
            self.ws = self.wb.create_sheet(self.sheet)
        else:
            self.ws = self.wb[self.sheet]

    # 连接db
    def connect_db(self, config):
        self.db = Db(config['ip'], config['port'], config['dbname'], config['user'], config['password'], config['echo_sql'])

    # 查询db
    def query_db(self, config):
        for var, sql in config.items():
            sql = replace_var(sql)
            df = self.db.query_dataFrame(sql)
            set_var(var, df)

    # 导出excel
    def export_excel(self, config):
        if isinstance(config, str):
            var_df = config
            config = {}
        else:
            var_df = config["df"]
        # 获得导出的变量
        val = self.get_var_DataFrame(var_df)
        if len(val) == 0:
            print(f"列表变量[{var_df}]为空, 不用导出excel")
            return

        # 导出
        # print(val)
        print(f'导出excel: {self.file}')
        if self.file.endswith('csv'):
            val.to_csv(self.file)
        else:
            if 'start' in config:
                startrow, startcol = self.split_bound(config['start'])
            else:
                startrow = 0
                startcol = 0
            writer = pd.ExcelWriter(self.file)
            val.to_excel(writer, self.sheet, index=False, startrow=startrow, startcol=startcol)

    # 获得指定变量, 并保证是DataFrame类型
    def get_var_DataFrame(self, var):
        val = get_var(var)
        # 检查类型
        if not isinstance(val, pd.DataFrame):
            if not isinstance(val, list):
                raise Exception(f"变量[{var}]值不是DataFrame或list: {val}")
            # list转DataFrame
            if len(val) == 0: # 空
                val = pd.DataFrame()
            else:
                fields = val[0].keys()
                val = pd.DataFrame(val, columns=fields)
            # 回写变量
            set_var(var, val)
        return val

    # df列变换
    def map_cols(self, cols, var_df):
        # 获得df
        df = self.get_var_DataFrame(var_df)
        # 构建df列变换器
        mapper = DfColMapper(df)
        # 逐列转换
        for col, expr in cols.items():
            mapper.map(col, expr)

    # 循环rows, 如 rows(1:3)
    # :param styles 每个迭代中要应用的样式
    # :param bound 范围
    def rows(self, styles, bound):
        self.do_for_styleable(styles, self.iterate_rows(bound), f"rows({bound})")

    # 循环cols, 如 cols(A:B)
    # :param styles 每个迭代中要应用的样式
    # :param bound 范围
    def cols(self, styles, bound):
        self.do_for_styleable(styles, self.iterate_cols(bound), f"cols({bound})")

    # 循环cells, 如 cells(A1:B3)
    # :param styles 每个迭代中要应用的样式
    # :param bound 范围
    def cells(self, styles, bound):
        self.do_for_styleable(styles, self.iterate_cells(bound), f"cells({bound})")

    # 循环对象应用样式
    def do_for_styleable(self, styles, objs, label):
        log.debug(f"-- Loop start: {label} -- ")
        for obj in objs:
            StyleableWrapper(obj).use_styles(styles)  # 应用样式
        log.debug(f"-- Loop finish: {label} -- ")

    # 迭代指定范围内的行
    def iterate_rows(self, bound):
        # return map(lambda row: self.ws.row_dimensions[row], self.build_row_range(bound))
        for row in self.build_row_range(bound):
            yield self.ws.row_dimensions[row]

    # 迭代指定范围内的列
    def iterate_cols(self, bound):
        for col in self.build_col_range(bound):
            yield self.ws.column_dimensions[col]

    # 构建行范围
    # :param bound 如 1:3
    def build_row_range(self, bound):
        # 1 两值
        if ':' in bound:
            # 分割开始值+结束值
            start, end = bound.split(':')
            return range(int(start), int(end) + 1)

        # 2 单值
        return [int(bound)]

    # 列顺序汇总
    Col_Idx = 'ABCDEFGHIGKLMNOPQRSTUVWXYZ'

    # 构建列范围
    # :param bound 如 A:B
    def build_col_range(self, bound):
        bound = bound.upper() # 转大写
        # 1 两值
        if ':' in bound:
            # 分割开始值+结束值
            start, end = bound.split(':')
            # 获得开始索引+结束索引
            start = Boot.Col_Idx.index(start)
            end = Boot.Col_Idx.index(end)
            # 构建yield迭代器
            for i in range(start, end + 1):
                yield Boot.Col_Idx[i]
            return

        # 2 单值
        return [bound]

    # 迭代指定范围内的单元格
    # https://blog.csdn.net/weixin_48668114/article/details/126444151
    def iterate_cells(self, bound):
        # 简单检查范围格式: 字母大小写都可以
        if re.match(r'[\w+\d:]+', bound) == None: # 匹配： 字母数字:
            if re.match(r'[\d,:]+', bound) == None:  # 匹配： 数字,:
                raise Exception('无效范围: ' + bound)

        # 1 纯数字: 如 1,2 或 1,2:3,4
        if ',' in bound:
            bs = self.split_bound(bound)
            if len(bs) == 4: # 1.1 范围: 起始行, 起始列, 结束行, 结束列
                for row in range(bs[0], bs[3] + 1):
                    for col in range(bs[1], bs[4] + 1):
                        yield self.ws.cell(row, col)
                return

            # 1.2 单个单元格: 起始行, 起始列
            yield self.ws.cell(bs[0], bs[1])
            return

        # 2 字母+数字
        # 2.1 范围 ws["A1:C3"], ws["A:C"], ws[1:3]
        if ':' in bound:
            for items in self.ws[bound]:
                for cell in items:
                    yield cell
            return

        # 2.2 单个单元格 ws["A1"]
        mat = re.match(r'\w+\d+', bound) # 匹配: 字母+数字
        if mat != None:
            yield self.ws[bound]
            return

        # 2.3 单列或单行 ws["A"], ws[1]
        for item in self.ws[bound]:
            yield item

    # 分割范围
    def split_bound(self, bound):
        bs = bound.split(':', 1)
        if len(bs) == 2:
            start = bs[0].split(',', 1)
            end = bs[1].split(',', 1)
            # 起始行, 起始列, 结束行, 结束列
            return [self.to_int(start[0]), self.to_int(start[1]), self.to_int(end[0], self.ws.max_row), self.to_int(end[1], self.ws.max_col)]

        start = bound.split(',', 1)
        # 起始行, 起始列
        return [self.to_int(start[0]), self.to_int(start[1])]

    # 转int, 有默认值(行号、列号从1开始)
    def to_int(self, str, default = 1):
        if str == None or str == '':
            return default

        return int(str)

    # 插入列
    # :param param 起始列,插入列数
    def insert_cols(self, param):
        idx, amount = param.split(',')
        self.ws.insert_cols(idx, amount)

    # 插入行
    # :param param 起始行,插入行数
    def insert_rows(self, param):
        idx, amount = param.split(',')
        self.ws.insert_rows(idx, amount)

    # 删除列
    # :param param 起始列,删除列数
    def remove_cols(self, param):
        idx, amount = param.split(',')
        self.ws.remove_cols(idx, amount)

    # 删除行
    # :param param 起始行,删除行数
    def remove_rows(self, param):
        idx, amount = param.split(',')
        self.ws.remove_rows(idx, amount)

    # 合并单元格
    def merge_cells(self, param):
        # self.wb.merge_cells(start_row=7, start_column=1, end_row=8, end_column=3)
        if re.match(r'\d+,\d+:\d+,\d+', param):
            start_row, start_column, end_row, end_column = param.replace(':', ',').split(',')
            self.wb.merge_cells(start_row=start_row, start_column=start_column, end_row=end_row, end_column=end_column)
            return

        # self.wb.merge_cells("C1:D2")
        return self.wb.merge_cells(param)

    # 取消合并单元格
    def unmerge_cells(self, param):
        # self.wb.unmerge_cells(start_row=7, start_column=1, end_row=8, end_column=3)
        if re.match(r'\d+,\d+:\d+,\d+', param):
            start_row, start_column, end_row, end_column = param.replace(':', ',').split(',')
            self.wb.unmerge_cells(start_row=start_row, start_column=start_column, end_row=end_row, end_column=end_column)
            return

        # self.wb.unmerge_cells("C1:D2")
        return self.wb.unmerge_cells(param)


# cli入口
def main():
    # 基于yaml的执行器
    boot = Boot()
    # 读元数据：author/version/description
    dir = os.path.dirname(__file__)
    meta = read_init_file_meta(dir + os.sep + '__init__.py')
    # 步骤配置的yaml
    step_files = parse_cmd('ExcelBoot', meta['version'])
    if len(step_files) == 0:
        raise Exception("Miss step config file or directory")
    try:
        # 执行yaml配置的步骤
        boot.run(step_files)
    except Exception as ex:
        log.error(f"Exception occurs: current step file is {boot.step_file}", exc_info = ex)
        raise ex

if __name__ == '__main__':
    main()